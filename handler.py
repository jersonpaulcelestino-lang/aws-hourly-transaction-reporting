import os
import logging
import boto3
import smtplib
from datetime import datetime, timedelta
from pymongo import MongoClient
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from openpyxl import Workbook, load_workbook
from io import BytesIO

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client("s3")

# ==========================================
# 1. DICCIONARIOS COMPLETOS
# ==========================================
from constants import currencyMap, reasonMap

# ==========================================
# 2. HELPERS
# ==========================================
def format_card(bin, endPan, brand):
    if not bin or not endPan:
        return None
    max_len = 15 if brand == 'AMERICAN_EXPRESS' else (14 if brand == 'DINERS' else 16)
    pad_len = max(0, max_len - len(str(bin)) - len(str(endPan)))
    return f"{bin}{'*'*pad_len}{endPan}"

def format_amount(amount):
    if not amount:
        return None
    s = str(amount)
    if len(s) > 2:
        return f"{s[:-2]}.{s[-2:]}"
    return f"0.{s}"

def get_time_range():
    now_utc = datetime.utcnow().replace(minute=0, second=0, microsecond=0)
    now_local = (now_utc - timedelta(hours=5))

    if now_local.hour == 0:
        # medianoche local → tomar 23:00–00:00 del día anterior
        start_local = now_local - timedelta(hours=1)
        end_local = now_local

        start_utc = start_local + timedelta(hours=5)
        end_utc = end_local + timedelta(hours=5)

        return start_utc, end_utc, start_local, end_local, True

    else:
        start_local = now_local - timedelta(hours=1)
        end_local = now_local

        start_utc = start_local + timedelta(hours=5)
        end_utc = end_local + timedelta(hours=5)

        return start_utc, end_utc, start_local, end_local, False

        
# ==========================================
# 3. MAIN
# ==========================================
def lambda_handler(event, context):

    mongo_uri = os.environ["db"]
    db_name = os.environ["DB_NAME"]
    bucket = os.environ["BUCKET_NAME"]
    bucketHora = os.environ["BUCKET_NAME_PER_HOUR"]

    smtp_server = os.environ["SMTP_SERVER"]
    smtp_port = int(os.environ["SMTP_PORT"])
    smtp_user = os.environ["SMTP_USER"]
    smtp_pass = os.environ["SMTP_PASS"]

    sender = os.environ["SENDER"]
    to_emails = [e.strip() for e in os.environ["TO"].split(",")]
    cc_emails = [e.strip() for e in os.environ["CC"].split(",")]
    bcc_list = [e.strip() for e in os.environ["BCC"].split(",")]

    start_date_utc, end_date_utc, start_date_local, end_date_local,is_midnight = get_time_range()

    client = MongoClient(mongo_uri)
    db = client[db_name]

    cursor = db.TDS_TRANSACTION.find({
        "issuerId": "XXX",
        "createdAt": {"$gte": start_date_utc, "$lt": end_date_utc}
    })

    results = []

    for tx in cursor:

        target_id = tx.get("threeDSServerTransID")
        acs_id = tx.get("acsTransID")

        ar = db.TDS_AREQ.find_one({"threeDSServerTransID": target_id}) or {}
        ares = db.TDS_ARES.find_one({"threeDSServerTransID": target_id}) or {}
        rreq = db.TDS_RREQ.find_one({"threeDSServerTransID": target_id}) or {}

        risk = db.TDS_RISK.find_one(
            {"acsTransID": acs_id} if acs_id else {"threeDSServerTransID": target_id}
        ) or {}

        # Resultado autenticación
        trans_status = tx.get("transStatus")
        if trans_status in ['Y','A']:
            auth_res = 'Autenticacion Exitosa'
        elif trans_status in ['N','R']:
            auth_res = 'Autenticacion Denegada'
        elif trans_status == 'U':
            auth_res = 'Autenticacion Fallida'
        elif trans_status == 'C':
            auth_res = 'Autenticacion Incompleta'
        else:
            auth_res = 'Otros'

        # Estado tarjeta
        c_status = (tx.get("rbaDataRequest") or {}).get("cardStatus")
        if c_status == '01':
            status_tarjeta = 'Afiliado'
        elif c_status == '02':
            status_tarjeta = 'Desafiliado'
        elif c_status == '03':
            status_tarjeta = 'Bloqueado'
        else:
            status_tarjeta = 'No enrolada'

        reason_code = rreq.get("transStatusReason") or ares.get("transStatusReason")

        results.append([
            format_card(tx.get("bin"), tx.get("endPan"), tx.get("brand")),
            tx.get("brand"),
            ar.get("acquirerMerchantID") or tx.get("acquirerMerchantID"),
            ar.get("merchantName"),
            currencyMap.get(ar.get("purchaseCurrency"), ar.get("purchaseCurrency")),
            format_amount(ar.get("purchaseAmount")),
            risk.get("valueRisk"),
            "otp",
            trans_status,
            auth_res,
            reason_code,
            reasonMap.get(reason_code),
            ares.get("eci") or rreq.get("eci"),
            status_tarjeta,
            ar.get("deviceChannel"),
            tx.get("beginDatetime") or tx.get("createdAt"),
            ares.get("messageVersion"),
            'S' if ares.get("acsDecConInd") == 'Y' else 'N',
            tx.get("endDatetime")
        ])

    # =========================
    # EXCEL
    # =========================
    headers = [
        "Numero de tarjeta","Marca","Codigo de Comercio","Comercio","Moneda","Monto",
        "Score RBA","Metodo de Autenticacion","Codigo de Resultado",
        "Resultado de Autenticacion","TransStatusReason","Descripcion de TransStatusReason",
        "ECI","Estado de Tarjeta","Canal de Compra","Fecha y hora de registro",
        "Version","Autenticacion desacoplada","Fecha y hora de autenticacion (Fin)"
    ]

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in results:
        ws.append(r)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    date_str_excel_por_hora = end_date_local.strftime("%Y%m%d")
    hour_exec_per_hour = end_date_local.strftime("%H")


    hourly_name = f"Reporte_Transacciones_{hour_exec_per_hour}_{date_str_excel_por_hora}.xlsx"
    
    s3.put_object(Bucket=bucket, Key=f"{bucketHora}/{hourly_name}", Body=buffer.getvalue())

    # =========================
    # ACUMULADO
    # =========================
    try:
        prev_hour = int(hour_exec_per_hour) - 1
        prev_file = f"Reporte_Transacciones_00-{prev_hour:02d}_{date_str_excel_por_hora}.xlsx"

        obj = s3.get_object(Bucket=bucket, Key=f"{prev_file}")
        prev_wb = load_workbook(BytesIO(obj["Body"].read()))
        prev_ws = prev_wb.active

        # Leer datos existentes
        existing_data = list(prev_ws.iter_rows(min_row=2, values_only=True))

        # Unir datos existentes con nuevos datos
        all_data = existing_data + results
        
        def parse_date(row):
            """
            Extrae la fecha de la fila si es un str YYYYMMDD, si no devuelve None.
            """
            value = row[15] # Columna 16 (índice 15)

            # El siguiente if no debería ejecutarse, pero es una validación por si acaso
            if isinstance(value, datetime):
                return value
            # El try-except se encarga de convertir el string a datetime, pero si falla devuelve datetime.min
            # Por lo tanto, la ordenación se hará correctamente.
            try:
                return datetime.strptime(str(value), "%d/%m/%Y %H:%M:%S")
            except:
                return datetime.min  # fallback

        # Ordenar por fecha (columna índice 15)
        # Solo tomamos filas donde la fecha sea válida (no None)
        # sorted : menor a mayor
        # reverse=True : mayor a menor
        all_data_sorted = sorted(all_data, key=parse_date, reverse=True)

        # Limpiar hoja (excepto header)
        prev_ws.delete_rows(2, prev_ws.max_row)

        for row in all_data_sorted:
            prev_ws.append(row)

        final_wb = prev_wb

    except Exception:
        final_wb = wb

    final_buffer = BytesIO()
    final_wb.save(final_buffer)
    final_buffer.seek(0)

    date_yesterday = start_date_local.strftime("%Y%m%d") 

    if is_midnight:
        final_name = f"Reporte_Transacciones_00-23_{date_yesterday}.xlsx"

        cuerpo_html = f"""
            <html>
            <body>
                <p>Estimados</p>
                <p>Se comparte el reporte de las transacciones acumuladas desde las 00 horas hasta las 23 horas, para su consideración</p>
                <p>Quedamos atentos a cualquier consulta adicional</p>
                <p>Saludos Cordiales,<br>Centro de Operaciones</p>
            </body>
            </html>
            """
    else:
        final_name = f"Reporte_Transacciones_00-{hour_exec_per_hour}_{date_str_excel_por_hora}.xlsx"
        cuerpo_html = f"""
        <html>
        <body>
            <p>Estimados</p>
            <p>Se comparte el reporte de las transacciones acumuladas desde las 00 horas hasta las {hour_exec_per_hour}, para su consideración</p>
            <p>Quedamos atentos a cualquier consulta adicional</p>
            <p>Saludos Cordiales,<br>Centro de Operaciones</p>
        </body>
        </html>
        """


    # Key=f"{final_name}" => Es la ruta y el nombre del archivo en S3
    # Body=final_buffer.getvalue() => Son los bytes del Excel que creamos en memoria
    s3.put_object(Bucket=bucket, Key=f"{final_name}", Body=final_buffer.getvalue())


    # =========================
    # EMAIL
    # =========================
    

    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = ",".join(to_emails)
    msg["Cc"] = ",".join(cc_emails)
    msg["Subject"] = f" REPORTE DE AUTENTICACIONES | CLIENTE FINANCIERO "

    msg.attach(MIMEText(cuerpo_html, "html"))

    attachment = MIMEApplication(final_buffer.getvalue())
    attachment.add_header("Content-Disposition", "attachment", filename=final_name)
    msg.attach(attachment)

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(sender, to_emails + cc_emails + bcc_list, msg.as_string())

    return {"status": "ok"} 